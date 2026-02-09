import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.template.loader import render_to_string
from weasyprint import HTML

# ==============================================================================
# 1. LAPORAN KEUANGAN (PROFIT & LOSS) - UPDATED
# ==============================================================================
def export_financial_report_to_excel(report_data):
    """
    Export Excel Laporan Laba Rugi dengan format standar Akuntansi.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Laba Rugi"

    # Styles
    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    money_fmt = '#,##0'
    center = Alignment(horizontal='center')

    # Title
    ws['A1'] = "LAPORAN LABA RUGI (PROFIT & LOSS)"
    ws['A1'].font = title_font
    ws['A1'].alignment = center
    ws.merge_cells('A1:C1')

    ws['A2'] = f"Periode: {report_data['start_date']} s/d {report_data['end_date']}"
    ws['A2'].alignment = center
    ws.merge_cells('A2:C2')

    # Data Rows
    # Format: (Label, Value, Is_Bold, Indent_Level)
    rows = [
        ("PENDAPATAN (REVENUE)", "", True, 0),
        ("Pendapatan Transaksi", report_data['total_income'], False, 1),
        ("", "", False, 0),
        ("BEBAN POKOK & OPERASIONAL", "", True, 0),
        ("(-) Pembelian Stok (HPP)", report_data['total_purchases'], False, 1),
        ("(-) Biaya Operasional", report_data['total_operational'], False, 1),
        ("Total Pengeluaran", report_data['total_expenses'], True, 1),
        ("", "", False, 0),
        ("LABA BERSIH (NET PROFIT)", report_data['net_profit'], True, 0)
    ]

    r = 4
    for label, val, is_bold, indent in rows:
        # Kolom Label
        cell_a = ws.cell(row=r, column=1, value=("   " * indent) + label)
        cell_a.font = bold if is_bold else Font(bold=False)
        
        # Kolom Nilai
        if val != "":
            cell_b = ws.cell(row=r, column=2, value=val)
            cell_b.number_format = money_fmt
            cell_b.font = bold if is_bold else Font(bold=False)
            
            # Warnai Laba Bersih
            if "LABA BERSIH" in label:
                cell_b.fill = PatternFill(start_color="C6EFCE" if val >= 0 else "FFC7CE", fill_type="solid")
        
        r += 1

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 25

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ==============================================================================
# 2. INVENTORY PDF - UPDATED
# ==============================================================================
def export_inventory_report_to_pdf(items_queryset):
    """
    Export Inventory ke PDF menggunakan WeasyPrint.
    items_queryset: Queryset yang sudah di-annotate (min_service_price, asset_value, dll)
    """
    context = {
        'items': items_queryset, 
        'generated_date': datetime.now()
    }
    html_string = render_to_string('reports/inventory_report_pdf.html', context)
    html = HTML(string=html_string)
    pdf_buffer = html.write_pdf()
    return io.BytesIO(pdf_buffer)

def export_inventory_report_to_excel(items_queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Stok Inventaris"

    # Header Title
    ws['A1'] = "LAPORAN STOK INVENTARIS (FULL DETAIL)"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Diunduh pada: {datetime.now().strftime('%d %B %Y %H:%M')}"

    # --- KOLOM FULL SESUAI PERMINTAAN ---
    headers = [
        "SKU", 
        "Nama Barang", 
        "Kategori", 
        "Tanggal Masuk (Join Date)",  # Kolom Baru
        "Stok", 
        "Status Stok", 
        "Harga Beli (Avg)", 
        "Harga Jual", 
        "Detail Jasa Pasang", 
        "Nilai Aset (Stok x Harga Beli)"
    ]
    
    ws.append([]) 
    ws.append(headers)

    # Styling Header
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Isi Data
    for item in items_queryset:
        # Status Stok
        if item.quantity == 0: status = "Habis"
        elif item.quantity <= item.reorder_threshold: status = "Menipis"
        else: status = "Aman"

        # Format Jasa Pasang (Digabung jadi text)
        services = [f"{s.vehicle_type}: {s.price:,.0f}" for s in item.service_prices.all()]
        services_str = ";\n".join(services) if services else "-"

        # Format Tanggal
        tgl_masuk = item.created_at.strftime('%d-%m-%Y') if item.created_at else "-"

        ws.append([
            item.sku or "-",
            item.name,
            item.category.name if item.category else "-",
            tgl_masuk,         # Data Tanggal
            item.quantity,
            status,
            item.buy_price,
            item.sell_price,
            services_str,      # Data Jasa Pasang Lengkap
            item.asset_value
        ])

    # Formatting Cell
    for row in range(5, ws.max_row + 1):
        # Format Rupiah
        for col in [7, 8, 10]: # Harga Beli, Jual, Aset
            ws.cell(row=row, column=col).number_format = '#,##0'
        
        # Wrap Text untuk Jasa Pasang
        ws.cell(row=row, column=9).alignment = Alignment(wrap_text=True)

    # Lebar Kolom
    ws.column_dimensions['B'].width = 35  # Nama
    ws.column_dimensions['D'].width = 18  # Tanggal
    ws.column_dimensions['I'].width = 45  # Jasa Pasang (Lebar)
    ws.column_dimensions['J'].width = 20  # Aset

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ==============================================================================
# 4. CUSTOMER REPORT EXCEL (Dari kode lama, dirapikan)
# ==============================================================================
def export_customer_report_to_excel(customers, start_date, end_date):
    """
    Export Laporan Pelanggan ke Excel.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Pelanggan"

    # Header
    ws['A1'] = "LAPORAN AKTIVITAS PELANGGAN"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Periode: {start_date} - {end_date}"

    headers = ['Nama Pelanggan', 'No. Telepon', 'Total Kunjungan', 'Total Belanja', 'Kunjungan Terakhir']
    ws.append([])
    ws.append(headers)

    # Style Header
    for col in range(1, 6):
        cell = ws.cell(row=4, column=col)
        cell.font = Font(bold=True)
        cell.border = Border(bottom=Side(style='thin'))

    for cust in customers:
        last_visit = cust.last_visit.strftime('%d-%m-%Y') if cust.last_visit else "-"
        ws.append([
            cust.name,
            cust.phone_number,
            cust.total_visits,
            cust.total_spending,
            last_visit
        ])

    # Format Currency Column D
    for row in range(5, ws.max_row + 1):
        ws.cell(row=row, column=4).number_format = '#,##0'

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ==============================================================================
# 5. EXPENSE BREAKDOWN EXCEL (Dari kode lama)
# ==============================================================================
def export_expense_breakdown_to_excel(report_data): 
    wb = Workbook()
    ws = wb.active
    ws.title = "Rincian Pengeluaran"
    
    ws['A1'] = "LAPORAN RINCIAN PENGELUARAN"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Periode: {report_data['start_date']} - {report_data['end_date']}"
    
    headers = ['Kategori', 'Frekuensi', 'Total Nominal', 'Persentase']
    ws.append([])
    ws.append(headers)
    
    # Style Header
    for col in range(1, 5):
        ws.cell(row=4, column=col).font = Font(bold=True)
    
    for row in report_data['breakdown']:
        ws.append([
            row['category'],
            row['count'],
            row['total'],
            f"{row['percent']}%"
        ])
        
    # Format Currency
    for r in range(5, ws.max_row + 1):
        ws.cell(row=r, column=3).number_format = '#,##0'

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['C'].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ==============================================================================
# 6. DEAD STOCK EXCEL (Dari kode lama)
# ==============================================================================
def export_dead_stock_to_excel(dead_stock_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dead Stock"
    
    ws['A1'] = "LAPORAN BARANG MATI (DEAD STOCK)"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Generated: {datetime.now().strftime('%d %b %Y')}"
    
    headers = ['Nama Barang', 'Stok Saat Ini', 'Terakhir Terjual', 'Hari Tidak Laku', 'Nilai Aset Mandeg']
    ws.append([])
    ws.append(headers)
    
    for col in range(1, 6):
        ws.cell(row=4, column=col).font = Font(bold=True)
        
    for item in dead_stock_data:
        item_name = item['item'].name
        qty = item['item'].quantity
        last = item['last_sale'].strftime('%d-%m-%Y') if item['last_sale'] else "Belum Pernah"
        days = item['days_inactive']
        asset = item['asset_value']
        
        ws.append([item_name, qty, last, days, asset])

    # Format Currency
    for r in range(5, ws.max_row + 1):
        ws.cell(row=r, column=5).number_format = '#,##0'

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['E'].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ==============================================================================
# 7. SALES REPORT EXCEL (Dari kode lama)
# ==============================================================================
def export_sales_report_to_excel(report_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    ws['A1'] = "LAPORAN PENJUALAN BARANG"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Periode: {report_data['start_date']} - {report_data['end_date']}"
    
    headers = ['Nama Barang', 'Kategori', 'Qty Terjual', 'Omzet (Revenue)', 'HPP', 'Profit', 'Margin %']
    ws.append([])
    ws.append(headers)
    
    for col in range(1, 8):
        ws.cell(row=4, column=col).font = Font(bold=True)
        
    for row in report_data['sales_data']:
        ws.append([
            row['name'],
            row['category'],
            row['qty'],
            row['revenue'],
            row['hpp'],
            row['profit'],
            f"{round(row['margin'], 1)}%"
        ])

    # Format Currency
    for r in range(5, ws.max_row + 1):
        ws.cell(row=r, column=4).number_format = '#,##0' # Revenue
        ws.cell(row=r, column=5).number_format = '#,##0' # HPP
        ws.cell(row=r, column=6).number_format = '#,##0' # Profit

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['F'].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer